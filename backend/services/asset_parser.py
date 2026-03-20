from pathlib import Path
from typing import Tuple


def parse_text_from_file(file_path: str) -> Tuple[str, str]:
    path = Path(file_path)
    ext = path.suffix.lower()
    if ext in {".txt", ".md", ".csv"}:
        return _read_text(path), "parsed"
    if ext == ".docx":
        return _read_docx(path), "parsed"
    if ext == ".pdf":
        return _read_pdf(path), "parsed"
    if ext == ".xlsx":
        return _read_xlsx(path), "parsed"
    if ext in {".png", ".jpg", ".jpeg", ".webp"}:
        return "", "image_uploaded_no_ocr"
    return "", "unsupported_type"


def _read_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="gbk", errors="ignore")


def _read_docx(path: Path) -> str:
    try:
        from docx import Document
    except Exception:
        return ""
    doc = Document(str(path))
    lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n".join(lines)


def _read_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        return ""
    reader = PdfReader(str(path))
    chunks = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            chunks.append(text.strip())
    return "\n\n".join(chunks)


def _read_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return ""

    workbook = load_workbook(path, read_only=True, data_only=True)
    chunks: list[str] = []
    for sheet in workbook.worksheets:
        chunks.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if cells:
                chunks.append(" | ".join(cells))
        chunks.append("")
    workbook.close()
    return "\n".join(chunks).strip()
